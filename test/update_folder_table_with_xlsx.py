# update_folder_table_with_xlsx.py

import openpyxl
import oracledb

# =========================
# 설정
# =========================
EXCEL_PATH = r"C:\Users\user\Downloads\update_folder.xlsx"   # 엑셀 파일명
START_ROW = 2                       # 1행이 헤더면 2, 헤더 없으면 1

DB_HOST = ""
DB_PORT = 1521
DB_SERVICE = "ORCLCDB"
DB_USER = ""
DB_PASSWORD = ""

# F열 값별 업데이트 값
PRESV_MAP = {
    "1": ("RCT-2-01", 1),
    "3": ("RCT-2-03", 3),
    "5": ("RCT-2-05", 5),
    "10": ("RCT-2-10", 10),
    "30": ("RCT-2-25", 25),
    "준영구": ("RCT-2-30", 30),
    "영구": ("RCT-2-40", 40),
}


def norm(value):
    """
    엑셀 값 정리
    - None 처리
    - 문자열 공백 제거
    - 2004.0 같은 숫자는 2004로 변환
    """
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def main():
    dsn = oracledb.makedsn(
        DB_HOST,
        DB_PORT,
        service_name=DB_SERVICE
    )

    conn = oracledb.connect(
        user=DB_USER,
        password=DB_PASSWORD,
        dsn=dsn
    )

    cursor = conn.cursor()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    updated_folder_ids = []

    try:
        for row_idx in range(START_ROW, ws.max_row + 1):
            # 엑셀 컬럼 순서
            # A열: tif.TSN_FOLDER_CREATE_ORG_CD
            # B열: tifo.TSN_FOLDER_CURKEEP_ORG_CD
            # C열: tif.TSN_FOLDER_TITLE
            # D열: tif.TNY_FOLDER_CREATE_YEAR
            # E열: tif.TNY_FOLDER_END_YEAR
            # F열: 보존기간 조건값

            create_org_cd = norm(ws.cell(row=row_idx, column=1).value)
            curkeep_org_cd = norm(ws.cell(row=row_idx, column=2).value)
            folder_title = norm(ws.cell(row=row_idx, column=3).value)
            create_year = norm(ws.cell(row=row_idx, column=4).value)
            end_year = norm(ws.cell(row=row_idx, column=5).value)
            f_value = norm(ws.cell(row=row_idx, column=6).value)

            # 완전 빈 줄 스킵
            if not any([
                create_org_cd,
                curkeep_org_cd,
                folder_title,
                create_year,
                end_year,
                f_value
            ]):
                continue

            # 필수값 누락 체크
            if not all([
                create_org_cd,
                curkeep_org_cd,
                folder_title,
                create_year,
                end_year,
                f_value
            ]):
                print(f"[SKIP] {row_idx}행 필수값 누락")
                print(
                    create_org_cd,
                    curkeep_org_cd,
                    folder_title,
                    create_year,
                    end_year,
                    f_value
                )
                continue

            if f_value not in PRESV_MAP:
                print(f"[SKIP] {row_idx}행 F열 값 매핑 없음: {f_value}")
                continue

            tsi_presv_term, tno_presv_term = PRESV_MAP[f_value]

            # 조인으로 업데이트 대상 TNO_FOLDER_ID 조회
            select_sql = """
                         SELECT tifm.TNO_FOLDER_ID
                         FROM tb_inf_folder_mng tifm
                                  JOIN tb_inf_folder tif
                                       ON tif.TNO_FOLDER_ID = tifm.TNO_FOLDER_ID
                                           AND tif.TNR_DEL = 0
                                  JOIN TB_INF_FOLDER_ORG tifo
                                       ON tifo.TNO_FOLDER_ID = tif.TNO_FOLDER_ID
                                           AND tifo.TNR_DEL = 0
                         WHERE tifm.TNR_DEL = 0
                           AND tif.TSN_FOLDER_CREATE_ORG_CD = :create_org_cd
                           AND tifo.TSN_FOLDER_CURKEEP_ORG_CD = :curkeep_org_cd
                           AND tif.TSN_FOLDER_TITLE = :folder_title
                           AND tif.TNY_FOLDER_CREATE_YEAR = :create_year
                           AND tif.TNY_FOLDER_END_YEAR = :end_year \
                         """

            params = {
                "create_org_cd": create_org_cd,
                "curkeep_org_cd": curkeep_org_cd,
                "folder_title": folder_title,
                "create_year": int(create_year),
                "end_year": int(end_year),
            }

            cursor.execute(select_sql, params)
            folder_ids = [row[0] for row in cursor.fetchall()]

            if not folder_ids:
                print(f"[NO MATCH] {row_idx}행 조건에 맞는 폴더 없음")
                print(
                    f"  create_org_cd={create_org_cd}, "
                    f"curkeep_org_cd={curkeep_org_cd}, "
                    f"title={folder_title}, "
                    f"create_year={create_year}, "
                    f"end_year={end_year}, "
                    f"f_value={f_value}"
                )
                continue

            # 조회된 TNO_FOLDER_ID 기준 업데이트
            update_sql = """
                         UPDATE tb_inf_folder_mng
                         SET TSI_FOLDER_PRESV_TERM = :tsi_presv_term,
                             TNO_FOLDER_PRESV_TERM = :tno_presv_term
                         WHERE TNO_FOLDER_ID = :tno_folder_id
                           AND TNR_DEL = 0 \
                         """

            for folder_id in folder_ids:
                cursor.execute(update_sql, {
                    "tsi_presv_term": tsi_presv_term,
                    "tno_presv_term": tno_presv_term,
                    "tno_folder_id": folder_id,
                })

                if cursor.rowcount > 0:
                    updated_folder_ids.append(folder_id)

                    print(
                        f"[UPDATE] {row_idx}행 "
                        f"/ TNO_FOLDER_ID={folder_id} "
                        f"/ TSI_FOLDER_PRESV_TERM={tsi_presv_term} "
                        f"/ TNO_FOLDER_PRESV_TERM={tno_presv_term}"
                    )
                else:
                    print(f"[UPDATE FAIL] {row_idx}행 / TNO_FOLDER_ID={folder_id}")

        conn.commit()

        print("\n==============================")
        print("업데이트 완료")
        print("==============================")
        print(f"총 업데이트 건수: {len(updated_folder_ids)}")

        print("\n업데이트된 TNO_FOLDER_ID 목록:")
        for folder_id in updated_folder_ids:
            print(folder_id)

    except Exception as e:
        conn.rollback()
        print("오류 발생. 전체 ROLLBACK 처리했습니다.")
        raise e

    finally:
        cursor.close()
        conn.close()


if __name__ == "__main__":
    main()